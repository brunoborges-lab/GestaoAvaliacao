import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook

# --- CONFIGURAÇÃO E CRITÉRIOS ---
CRITERIOS = {
    "Ferramentas": ["Transporta as ferramentas...", "Opera com a ferramenta...", "Coloca-se do lado certo...", "Efectua cominucação...", "Protege a(s) vítima(s)..."],
    "Equipamentos": ["Escolhe  equipamento...", "Transporta  e opera...", "Opera corretamente grupo...", "Opera corretamente estabilização...", "Opera corretamente pneumático..."],
    "Estabilização": ["Sinaliza e delimita...", "Estabiliza o(s) veículo(s)...", "Controla estabilização...", "Efetua limpeza...", "Aplica as proteções..."]
}

if 'db' not in st.session_state: st.session_state.db = {}

st.title("Gestor de Avaliaçãol")

# --- SIDEBAR: CARREGAMENTO DE MODELOS ---
with st.sidebar:
    st.header("Configuração")
    f_import = st.file_uploader("1. Importação (Nomes Coluna K)", type=["xlsx"])
    f_pratica = st.file_uploader("2. Modelo Prática (.xlsm)", type=["xlsm"])
    f_final = st.file_uploader("3. Modelo Avaliação Final (.xlsx)", type=["xlsx"])

if f_import and f_pratica and f_final:
    # Extração de Nomes
    df_nomes = pd.read_excel(f_import, skiprows=12, usecols="K").dropna()
    lista_nomes = df_nomes.iloc[:, 0].astype(str).tolist()
    
    formando = st.selectbox("Formando a avaliar:", lista_nomes)

    # --- FORMULÁRIO DE LANÇAMENTO ---
    with st.form("lancamento_notas"):
        nt = st.number_input("Nota Teórica", 0.0, 20.0, 10.0)
        
        st.write("### Notas de Avaliação Prática (1, 3, 5)")
        c1, c2, c3 = st.columns(3)
        notas_p = {}
        for i, (cat, itens) in enumerate(CRITERIOS.items()):
            with [c1, c2, c3][i]:
                st.markdown(f"**{cat}**")
                for idx, item in enumerate(itens):
                    notas_p[f"{cat}_{idx}"] = st.radio(f"{item[:30]}...", [1, 3, 5], index=1, key=f"{formando}_{cat}_{idx}")

        if st.form_submit_button("💾 Guardar Dados do Formando"):
            # Cálculo das médias por categoria (0 a 20)
            m_ferr = (sum([notas_p[f"Ferramentas_{i}"] for i in range(5)])/25)*20
            m_equip = (sum([notas_p[f"Equipamentos_{i}"] for i in range(5)])/25)*20
            m_estab = (sum([notas_p[f"Estabilização_{i}"] for i in range(5)])/25)*20
            
            st.session_state.db[formando] = {
                "teorica": nt, "pratica_detalhe": notas_p,
                "m_ferr": m_ferr, "m_equip": m_equip, "m_estab": m_estab
            }
            st.success(f"Avaliação de {formando} registada!")

    # --- BOTÃO DE EXPORTAÇÃO ---
    if st.session_state.db:
        st.divider()
        if st.button("🚀 Gerar Todos os Ficheiros Preenchidos"):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                
                # 1. PROCESSAR FICHEIROS DE AVALIAÇÃO PRÁTICA (Individuais)
                f_pratica.seek(0)
                template_p = f_pratica.read()
                for nome, dados in st.session_state.db.items():
                    wb_p = load_workbook(io.BytesIO(template_p), keep_vba=True)
                    ws_p = wb_p.active
                    ws_p['C6'] = nome  # Nome do Formando
                    
                    # Marcar X (Exemplo Colunas AH, AI, AJ)
                    col_map = {1: 34, 3: 35, 5: 36}
                    # (Lógica de busca por texto omitida aqui por brevidade, mas segue o padrão anterior)
                    
                    p_out = io.BytesIO()
                    wb_p.save(p_out)
                    zf.writestr(f"Pratica_{nome}.xlsm", p_out.getvalue())

                # 2. PROCESSAR FICHEIRO DE AVALIAÇÃO FINAL (Único)
                f_final.seek(0)
                wb_f = load_workbook(io.BytesIO(f_final.read()))
                ws_f = wb_f.active
                # Colunas: Teórica (U/21), Ferr (AF/32), Equip (AP/42), Estab (AZ/52)
                for row in ws_f.iter_rows(min_row=12, max_row=100):
                    nome_excel = str(row[2].value).strip() # Coluna C
                    if nome_excel in st.session_state.db:
                        d = st.session_state.db[nome_excel]
                        ws_f.cell(row=row[0].row, column=21).value = d['teorica']
                        ws_f.cell(row=row[0].row, column=32).value = d['m_ferr']
                        ws_f.cell(row=row[0].row, column=42).value = d['m_equip']
                        ws_f.cell(row=row[0].row, column=52).value = d['m_estab']
                
                f_out = io.BytesIO()
                wb_f.save(f_out)
                zf.writestr("Avaliacao_Final_UFCD.xlsx", f_out.getvalue())

            st.download_button("📥 Descarregar Dossier Completo (ZIP)", zip_buffer.getvalue(), "Dossier_UFCD.zip")
